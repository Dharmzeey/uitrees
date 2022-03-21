from openpyxl import load_workbook

from trees.models import Tree, Order, Family, Authority, ConservationStatus, SpecieStatus
import time


def run():
    start = time.perf_counter()
    wb = load_workbook('excel/Exceldb.xlsx')
    ws = wb.active

    m_row = ws.max_row
    m_col = ws.max_column
    Tree.objects.all().delete()

    for row in range(2, 125 + 1):
        sn = ws.cell(row=row, column=1).value
        lk = ws.cell(row=row, column=2).value
        ln = ws.cell(row=row, column=3).value
        od = ws.cell(row=row, column=4).value
        fm = ws.cell(row=row, column=5).value
        at = ws.cell(row=row, column=6).value
        ss = ws.cell(row=row, column=7).value
        cs = ws.cell(row=row, column=8).value
        cn = ws.cell(row=row, column=9).value
        gs = ws.cell(row=row, column=10).value
        # for col in range(1, m_col + 1):
        #     cell_obj = ws.cell(row=row, column=col)

        get_od, created = Order.objects.get_or_create(tree_order=od)
        get_fm, created = Family.objects.get_or_create(tree_family=fm)
        get_at, created = Authority.objects.get_or_create(tree_authority=at)
        get_cs, created = ConservationStatus.objects.get_or_create(tree_status=cs)
        get_ss, created = SpecieStatus.objects.get_or_create(specie_status=ss)

        tr = Tree(
            scientific_name=sn,
            genus_specie=gs,
            authority=get_at,
            common_name=cn,
            local_name=ln,
            conservation_status=get_cs,
            specie_status=get_ss,
            tree_order=get_od,
            tree_family=get_fm,
            tree_description='',
            reference=lk,
        )
        tr.save()

    end = time.perf_counter()
    print(end - start)
